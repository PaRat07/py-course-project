from typing import Any
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from mylib.common.exception import MylibException

class SerializingException(MylibException): ...
class DeserializingException(MylibException): ...

def xlsx_to_matrix(path: str) -> list[list[Any]]:
    try:
        wb = load_workbook(path)
        ws = wb.active
        if ws is None:
            raise DeserializingException(f"No active worksheet found in {path}")

        matrix = []
        for row in ws.iter_rows(values_only=True):
            matrix.append(list(row))

        return matrix
    except Exception as e:
        raise DeserializingException("Failed to read Excel file") from e

def matrix_to_xlsx(path: str, matrix: list[list[Any]]) -> None:
    try:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise SerializingException("Failed to create worksheet")

        for row in matrix:
            ws.append(row)

        wb.save(path)
    except Exception as e:
        raise SerializingException(f"Failed to save Excel file") from e

